import yaml
import os
from openpyxl import load_workbook
algo_dir = '/ai/base/G-FIR/src/algos/'
algos = ['GreedHash', 'csq', 'DPgQ', 'GPQ']

def fill_result(algo, result, sheet_metric):
    dest = 'G-FIR exps.xlsx'
    #Open an xlsx for reading
    wb = load_workbook(filename = dest)
    #Get the current Active Sheet
    # ws = wb.get_active_sheet()
    #You can also select a particular sheet
    #based on sheet name
    ws = wb["DomainNet-"+sheet_metric]
    #Open the csv file
    hold_out_domain = None
    for row_id in range(2, 30):
        print('=============')
        row = str(row_id + 1)
        print('row = ', row)
        row = ws[row]
        # print(row)

        id = 0
        for col in row:

            if id==1 and col.value is not None:
                hold_out_domain = col.value.replace('-', '')

            if id == 2 and col.value is not None:
                # print(id,hold_out_domain, col.value)
                method = col.value
                print(id, hold_out_domain, method)

                if method.lower() == algo.lower():
                    # print(result[hold_out_domain])
                    j = 2
                    for bit in [16, 32, 64]:
                        ws.cell(row=row_id+1, column=id+j).value = result[hold_out_domain][bit][0][sheet_metric]
                        j += 1

                    for bit in [16, 32, 64]:
                        ws.cell(row=row_id+1, column=id+j).value = result[hold_out_domain][bit][1][sheet_metric]
                        j += 1
            print(id, col.value)
            id += 1

    wb.save(dest)

for algo in algos:
    result_filepath = os.path.join(algo_dir, algo, 'domainnet_result.yaml')
    with open(result_filepath, "r") as stream:
        try:
            result = yaml.safe_load(stream)
            fill_result(algo, result, 'aps@200')
            fill_result(algo, result, 'prec@200')
            fill_result(algo, result, 'recall@200')
        except yaml.YAMLError as exc:
            print(exc)
